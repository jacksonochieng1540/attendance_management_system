"""
Export helpers.

Both functions take an Attendance queryset (already filtered by date range,
department, employee, status, etc.) and return an HttpResponse containing
a downloadable file — Excel via openpyxl, PDF via reportlab.
"""

from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

HEADERS = ['Date', 'Employee ID', 'Name', 'Department', 'Status', 'Check In', 'Check Out', 'Remarks']


def _row_for(record):
    return [
        record.date.strftime('%d-%m-%Y'),
        record.employee.employee_id,
        record.employee.full_name,
        record.employee.department.name if record.employee.department else '-',
        record.get_status_display(),
        record.check_in_time.strftime('%H:%M') if record.check_in_time else '-',
        record.check_out_time.strftime('%H:%M') if record.check_out_time else '-',
        record.remarks or '-',
    ]


def export_attendance_to_excel(queryset, filename='attendance_report.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for record in queryset.select_related('employee', 'employee__department'):
        ws.append(_row_for(record))

    # Auto-fit column widths roughly
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = length + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_attendance_to_pdf(queryset, filename='attendance_report.pdf', title='Attendance Report'):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()

    elements = [Paragraph(title, styles['Title']), Spacer(1, 12)]

    data = [HEADERS]
    for record in queryset.select_related('employee', 'employee__department'):
        data.append(_row_for(record))

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
